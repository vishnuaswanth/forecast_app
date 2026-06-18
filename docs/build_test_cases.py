from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Palette ────────────────────────────────────────────────────────────────────
HDR_BG   = "1F3864"   # dark navy
HDR_FG   = "FFFFFF"
ROW_ALT  = "EBF3FB"   # light blue alt row
COL_CRIT = "FF0000"   # Critical  – red
COL_HIGH = "FF6600"   # High      – orange
COL_MED  = "0070C0"   # Medium    – blue
COL_LOW  = "00B050"   # Low       – green
STS_BG   = "F2F2F2"   # status col bg

TAB_COLORS = {
    "Summary":     "000000",
    "Auth":        "7030A0",
    "Upload":      "FF6600",
    "DataView":    "008080",
    "ManagerView": "002060",
    "ExecMonitor": "404040",
    "EditView":    "C00000",
    "Config":      "375623",
    "Chat":        "4F00A0",
    "Cache":       "7B3F00",
}

PRIORITY_COLORS = {
    "Critical": COL_CRIT,
    "High":     COL_HIGH,
    "Medium":   COL_MED,
    "Low":      COL_LOW,
}

COLS = ["TC ID", "Domain", "Feature", "Test Case Title",
        "Preconditions", "Test Steps", "Expected Result",
        "Acceptance Criteria", "Priority", "Status"]
COL_W = [14, 18, 20, 38, 32, 48, 38, 44, 12, 14]

STATUS_DV  = '"Not Run,Pass,Fail,Blocked,In Progress"'
PRIORITY_DV = '"Critical,High,Medium,Low"'


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def make_header(ws):
    hdr_font  = Font(name="Arial", bold=True, color=HDR_FG, size=11)
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (col, w) in enumerate(zip(COLS, COL_W), start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def add_row(ws, row_num, values):
    is_alt = (row_num % 2 == 0)
    bg = ROW_ALT if is_alt else "FFFFFF"
    pri = values[8] if len(values) > 8 else ""
    pri_color = PRIORITY_COLORS.get(pri, "000000")

    for i, val in enumerate(values, start=1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.font = Font(name="Arial", size=10,
                      color=pri_color if i == 9 else "000000",
                      bold=(i == 9))
        c.border = thin_border()
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if i == 10:  # Status
            c.fill = PatternFill("solid", fgColor=STS_BG)
        else:
            c.fill = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[row_num].height = 70


def add_dv(ws, num_rows):
    dv_status = DataValidation(type="list", formula1=STATUS_DV, allow_blank=False,
                               showErrorMessage=True, errorTitle="Invalid",
                               error="Choose from: Not Run, Pass, Fail, Blocked, In Progress")
    dv_status.sqref = f"J2:J{num_rows+1}"
    ws.add_data_validation(dv_status)

    dv_pri = DataValidation(type="list", formula1=PRIORITY_DV, allow_blank=False,
                            showErrorMessage=True, errorTitle="Invalid",
                            error="Choose: Critical, High, Medium, Low")
    dv_pri.sqref = f"I2:I{num_rows+1}"
    ws.add_data_validation(dv_pri)


def make_sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLORS.get(name, "000000")
    make_header(ws)
    return ws


def write_rows(ws, rows):
    for i, r in enumerate(rows, start=2):
        add_row(ws, i, r)
    add_dv(ws, len(rows))


# ── AUTH ───────────────────────────────────────────────────────────────────────
AUTH_ROWS = [
    ["TC-AUTH-001","Authentication","Login","Valid login with correct credentials","App is running; valid user exists in LDAP","1. Navigate to /centene_forecasting/\n2. Enter valid portal_id and password\n3. Click Login","User is redirected to timezone selection page; session is created","User can log in with valid LDAP credentials; session is established","Critical","Not Run"],
    ["TC-AUTH-002","Authentication","Login","Login with invalid credentials","App is running","1. Navigate to /centene_forecasting/\n2. Enter invalid portal_id or password\n3. Click Login","Error message displayed; user stays on login page","Invalid credentials are rejected with a clear error message","Critical","Not Run"],
    ["TC-AUTH-003","Authentication","Login","Login with empty fields","App is running","1. Navigate to /centene_forecasting/\n2. Leave both fields blank\n3. Click Login","Form validation error shown; login not attempted","Required field validation prevents empty submissions","High","Not Run"],
    ["TC-AUTH-004","Authentication","Timezone","Timezone selection saved to session","User is logged in (just after login)","1. Select a timezone from the dropdown (e.g. CST)\n2. Click Confirm/Save","User redirected to upload page; timezone stored in session","Selected timezone is persisted and used for all subsequent time displays","High","Not Run"],
    ["TC-AUTH-005","Authentication","Timezone","Default timezone pre-selected","User logged in and on timezone page","1. Navigate to timezone selection page\n2. Observe the default selection","A default timezone (IST) is pre-selected","A sensible default timezone is shown so the user doesn't have to pick","Medium","Not Run"],
    ["TC-AUTH-006","Authentication","Logout","Logout clears session","User is logged in","1. Click the Logout link/button","User redirected to login page; session is cleared; back button does not restore session","Logout fully invalidates the session","Critical","Not Run"],
    ["TC-AUTH-007","Authorization","Permissions","User with only 'view' permission cannot access upload","User logged in with view-only role","1. Navigate to /centene_forecasting/upload_view/","403 error or redirect to fallback page","Users without 'add' permission cannot access upload functionality","Critical","Not Run"],
    ["TC-AUTH-008","Authorization","Permissions","User with 'add' permission can access upload","User logged in with ADMIN/MANAGER role","1. Navigate to /centene_forecasting/upload_view/","Upload page loads successfully","Users with correct permissions can access restricted pages","High","Not Run"],
    ["TC-AUTH-009","Authorization","Permissions","Unauthenticated access redirects to login","No user session","1. Directly navigate to /centene_forecasting/data_view/","Redirected to login page","All protected pages require authentication","Critical","Not Run"],
    ["TC-AUTH-010","Authorization","Fallback","PermissionFallbackMiddleware redirects 403","User logged in but lacks specific permission","1. Try to access a page without permission","Friendly fallback page or redirect instead of raw 403","Authenticated users get a graceful error instead of raw HTTP 403","High","Not Run"],
    ["TC-AUTH-011","Authentication","Session","Session persists across page navigations","User is logged in","1. Navigate between multiple pages (data view, manager view, etc.)","User remains logged in without re-authentication","Session cookie maintains user identity throughout the session","Medium","Not Run"],
    ["TC-AUTH-012","Authentication","Login","Login page inaccessible when already logged in","User is already logged in","1. Navigate directly to /centene_forecasting/","Redirected away from login page (to upload or data view)","Already-authenticated users are not shown the login form again","Low","Not Run"],
]

# ── UPLOAD ─────────────────────────────────────────────────────────────────────
UPLOAD_ROWS = [
    ["TC-UPL-001","Upload","File Upload","Upload valid forecast CSV file","User logged in with 'add' permission; FastAPI backend running","1. Navigate to /centene_forecasting/upload_view/\n2. Select file type 'Forecast'\n3. Choose a valid .csv file\n4. Click Upload","Upload starts; progress bar appears; success message shown on completion","A valid CSV forecast file is accepted, processed, and confirmed with a success notification","Critical","Not Run"],
    ["TC-UPL-002","Upload","File Upload","Upload valid roster Excel file","User logged in with 'add' permission","1. Select file type 'Roster'\n2. Choose a valid .xlsx file\n3. Click Upload","Upload processes successfully; success notification shown","A valid XLSX roster file is accepted and processed","Critical","Not Run"],
    ["TC-UPL-003","Upload","File Upload","Upload invalid file type (e.g. .pdf)","User logged in with 'add' permission","1. Select any file type\n2. Choose a .pdf file\n3. Click Upload","Error message shown; upload not started","Only CSV and XLSX files are accepted; all other formats are rejected with a clear error","Critical","Not Run"],
    ["TC-UPL-004","Upload","File Upload","Upload empty file","User logged in with 'add' permission","1. Select any file type\n2. Choose a 0-byte CSV file\n3. Click Upload","Error message shown; upload rejected","Empty files are rejected before processing","High","Not Run"],
    ["TC-UPL-005","Upload","File Upload","Upload altered forecast file","User logged in with 'add' permission","1. Select file type 'Altered Forecast'\n2. Choose a valid .xlsx file\n3. Click Upload","Upload processes successfully","Altered forecast file type is accepted and processed correctly","High","Not Run"],
    ["TC-UPL-006","Upload","File Upload","Upload prod team roster file","User logged in with 'add' permission","1. Select file type 'Prod Team Roster'\n2. Choose a valid .xlsx file\n3. Click Upload","Upload processes successfully","Prod team roster file type is accepted and processed correctly","High","Not Run"],
    ["TC-UPL-007","Upload","Progress","Real-time progress tracking via WebSocket","User logged in; file upload initiated","1. Upload a large file\n2. Observe the progress indicator","Progress percentage updates in real-time (0%→100%) without page refresh","Upload progress is communicated to the user in real-time via WebSocket","Critical","Not Run"],
    ["TC-UPL-008","Upload","Progress","Check upload progress API","User logged in; upload in progress","1. While a file is uploading, call /centene_forecasting/check_upload_progress/","Returns current progress percentage as JSON","Progress API returns a numeric percentage reflecting actual processing state","High","Not Run"],
    ["TC-UPL-009","Upload","Cache","Forecast upload clears forecast cache","Data already cached from a prior load","1. Load data view to populate cache\n2. Upload a new forecast file\n3. Reload data view","Data view reflects new data (cache was cleared)","Uploading a forecast file invalidates stale cached data","Critical","Not Run"],
    ["TC-UPL-010","Upload","Cache","Roster upload clears roster cache","Roster data cached","1. Load roster data to populate cache\n2. Upload a new roster file\n3. Reload roster data","Roster data reflects new upload","Uploading a roster file invalidates stale roster cache","High","Not Run"],
    ["TC-UPL-011","Upload","History","Upload history is displayed","Previous uploads exist","1. Navigate to upload page","A list of recent uploads with status is shown","Upload history is visible so users can track past uploads","Medium","Not Run"],
    ["TC-UPL-012","Upload","Permissions","User without 'add' permission cannot upload","User logged in as viewer","1. Navigate to /centene_forecasting/upload_view/","Access denied (403 or redirect)","Upload functionality is restricted to users with 'add' permission","High","Not Run"],
    ["TC-UPL-013","Upload","File Upload","Upload without selecting file type","User logged in with 'add' permission","1. Leave file type unselected\n2. Choose a file\n3. Click Upload","Validation error shown","File type is required; upload is blocked if not selected","Medium","Not Run"],
    ["TC-UPL-014","Upload","File Upload","Upload without selecting a file","User logged in; file type selected","1. Select file type\n2. Do not choose a file\n3. Click Upload","Validation error shown","File selection is required; upload is blocked if no file chosen","Low","Not Run"],
]

# ── DATA VIEW ─────────────────────────────────────────────────────────────────
DV_ROWS = [
    ["TC-DV-001","Data View","Page Load","Data view page loads successfully","User logged in with 'view' permission","1. Navigate to /centene_forecasting/data_view/","Page loads with filter dropdowns and empty table","Data view page renders with all UI elements present","High","Not Run"],
    ["TC-DV-002","Data View","Cascade Filter","Year dropdown populates on load","User on data view page","1. Observe the Year dropdown on page load","Year dropdown is populated with available years","Available forecast years are loaded automatically when page opens","Critical","Not Run"],
    ["TC-DV-003","Data View","Cascade Filter","Month dropdown populates after selecting year","Year selected","1. Select a year from the Year dropdown","Month dropdown updates with months available for that year","Months are filtered to only those with data for the selected year","Critical","Not Run"],
    ["TC-DV-004","Data View","Cascade Filter","Platform dropdown cascades from month","Year and month selected","1. Select year\n2. Select month","Platform dropdown populates with available platforms","Platforms cascade correctly based on year+month selection","Critical","Not Run"],
    ["TC-DV-005","Data View","Cascade Filter","Market dropdown cascades from platform","Year, month, platform selected","1. Select year, month, platform","Market dropdown populates","Markets cascade correctly based on platform selection","High","Not Run"],
    ["TC-DV-006","Data View","Cascade Filter","Locality dropdown cascades from market","Year, month, platform, market selected","1. Select year, month, platform, market","Locality dropdown populates","Localities cascade correctly","High","Not Run"],
    ["TC-DV-007","Data View","Cascade Filter","WorkType dropdown populates","Filters selected","1. Complete all prior cascade steps","WorkType dropdown populates","Work types load correctly as the final cascade level","Medium","Not Run"],
    ["TC-DV-008","Data View","Cascade Filter","Changing a parent filter resets child dropdowns","Year selected, then month selected","1. Select year and month\n2. Change the year","Month and all subsequent dropdowns are cleared/reset","Changing a parent filter clears all dependent dropdowns to prevent stale selections","Critical","Not Run"],
    ["TC-DV-009","Data View","Data Table","Forecast data table loads with filters","All cascading filters selected","1. Apply all filters\n2. Click Load/Apply","Forecast table populates with matching records","Forecast data table returns records matching the selected filter combination","Critical","Not Run"],
    ["TC-DV-010","Data View","Data Table","Roster data table loads","Filters applied","1. Select roster tab\n2. Apply filters","Roster table populates","Roster data table returns records for the given filter context","High","Not Run"],
    ["TC-DV-011","Data View","Data Table","Actuals data table loads","Filters applied","1. Select actuals tab\n2. Apply filters","Actuals table populates","Actuals data is displayed correctly","High","Not Run"],
    ["TC-DV-012","Data View","Data Table","Pagination works correctly","Data loaded with many rows","1. Load data with many records\n2. Navigate to next page","Next page of records loads correctly","Server-side pagination works; each page change fetches the correct records","High","Not Run"],
    ["TC-DV-013","Data View","Data Table","Empty result handled gracefully","Filters selected with no matching data","1. Apply filters for a combination with no data","Table shows 'No data' message (not an error)","Empty result sets display a user-friendly 'no records found' message","Medium","Not Run"],
    ["TC-DV-014","Data View","Download","Download filtered data as Excel","Filters applied; data visible in table","1. Click the Download button","An Excel file is downloaded with the filtered data","Download produces a valid Excel file containing the currently filtered dataset","High","Not Run"],
    ["TC-DV-015","Data View","Filter Validation","Invalid year parameter rejected","User attempts direct URL manipulation","1. Call /centene_forecasting/forecast/months/?year=abc","400 error with meaningful message","Non-numeric or out-of-range year values are rejected by the API","Medium","Not Run"],
    ["TC-DV-016","Data View","Cache","Cascade filter results are cached","First request made","1. Select year (triggers API call)\n2. Select same year again quickly","Second selection is served from cache (faster response)","Filter cascade results are cached to avoid redundant API calls","Medium","Not Run"],
    ["TC-DV-017","Data View","Cache","Cache is invalidated after file upload","Data cached; new file uploaded","1. Load data to warm cache\n2. Upload new forecast file\n3. Re-run the same query","Fresh data is returned (not the previously cached result)","Cache invalidation ensures data freshness after uploads","High","Not Run"],
    ["TC-DV-018","Data View","Permissions","User without 'view' permission is blocked","User without any permission","1. Navigate to /centene_forecasting/data_view/","Access denied","Data view page requires at minimum 'view' permission","Medium","Not Run"],
]

# ── MANAGER VIEW ──────────────────────────────────────────────────────────────
MV_ROWS = [
    ["TC-MV-001","Manager View","Page Load","Manager view page loads successfully","User logged in with 'view' permission","1. Navigate to /centene_forecasting/manager-view/","Dashboard page loads with filter options and empty KPI cards","Manager view page renders correctly with all sections visible","High","Not Run"],
    ["TC-MV-002","Manager View","Filter","Report month dropdown populated","User on manager view page","1. Observe the Report Month dropdown","Dropdown shows available YYYY-MM values (last 6+ months)","Report month options are available for selection","High","Not Run"],
    ["TC-MV-003","Manager View","Dashboard Data","Dashboard data loads for selected month","Report month selected","1. Select a report month\n2. Click Load/Apply","Hierarchical category table populates with 6-month forecast data","Selecting a report month loads category-level forecast data for that period","Critical","Not Run"],
    ["TC-MV-004","Manager View","Dashboard Data","Category filter narrows data","Report month selected","1. Select a report month\n2. Select a specific category\n3. Load","Table shows only rows for the selected category","Category filter correctly scopes the dashboard data","High","Not Run"],
    ["TC-MV-005","Manager View","KPIs","KPI cards display correct values","Report month selected and data loaded","1. Load dashboard data\n2. Observe KPI summary cards","FTE, CPH, Capacity, Gap values are shown in the header cards","KPI summary cards show accurate calculated metrics for the selected period","Critical","Not Run"],
    ["TC-MV-006","Manager View","KPIs","KPI month index uses config value","Config set KPI_MONTH_INDEX = 1","1. Load dashboard\n2. Verify which month's data is used for KPIs","KPIs reflect data from the second month (index 1) in the 6-month window","KPI cards use the month defined by KPI_MONTH_INDEX config","Medium","Not Run"],
    ["TC-MV-007","Manager View","Hierarchy","Category hierarchy displays correctly","Multi-level category data available","1. Load dashboard\n2. Expand a parent category","Child categories appear nested under parent","Hierarchical category structure is displayed with expand/collapse","High","Not Run"],
    ["TC-MV-008","Manager View","Hierarchy","Max hierarchy depth is respected","Deep category tree available","1. Load data with deep category nesting","Categories only displayed up to MAX_HIERARCHY_DEPTH (6) levels","Hierarchy depth is capped at configuration limit to prevent runaway nesting","Medium","Not Run"],
    ["TC-MV-009","Manager View","API","Data API returns correct structure","Report month selected","1. Call /centene_forecasting/api/manager-view/data/?report_month=2025-01","JSON with success=true, months array, categories array","API returns well-structured response with all expected fields","High","Not Run"],
    ["TC-MV-010","Manager View","Validation","Invalid report_month format rejected","User manipulates URL parameters","1. Call data API with report_month=January (wrong format)","400 error returned","Only YYYY-MM format is accepted for report_month","High","Not Run"],
    ["TC-MV-011","Manager View","Validation","Missing report_month returns error","User calls API without required param","1. Call /centene_forecasting/api/manager-view/data/ with no params","400 error with helpful message","report_month is required; missing it returns a clear validation error","Medium","Not Run"],
    ["TC-MV-012","Manager View","Data","6 months of data displayed in table","Data available for 6 months","1. Load manager view with valid report month","Table columns show 6 monthly forecast columns","Dashboard always shows exactly 6 forecast months per configuration","Critical","Not Run"],
    ["TC-MV-013","Manager View","Permissions","Viewer can access manager view","User with 'view' permission","1. Navigate to manager view as a viewer","Page accessible","View permission is sufficient to access the manager dashboard","Low","Not Run"],
    ["TC-MV-014","Manager View","Permissions","Unauthenticated user blocked","No session","1. Navigate to /centene_forecasting/manager-view/","Redirected to login","Manager view requires authentication","Medium","Not Run"],
]

# ── EXEC MONITOR ──────────────────────────────────────────────────────────────
EM_ROWS = [
    ["TC-EM-001","Exec Monitor","Page Load","Execution monitoring page loads","User logged in with 'view' permission","1. Navigate to /centene_forecasting/execution-monitoring/","Page loads with execution list and KPI section","Execution monitoring page renders with all sections visible","High","Not Run"],
    ["TC-EM-002","Exec Monitor","Execution List","Execution list loads with records","Executions exist in the system","1. Load the execution monitoring page","List of executions shown with status, timestamp, and work type","Execution history is displayed in paginated list format","High","Not Run"],
    ["TC-EM-003","Exec Monitor","Execution List","Execution list filtered by status","Executions with various statuses exist","1. Apply Status filter (e.g. FAILED)\n2. Submit","Only FAILED executions are shown","Status filter correctly narrows the execution list","High","Not Run"],
    ["TC-EM-004","Exec Monitor","Execution List","Execution list filtered by month/year","Executions for multiple periods exist","1. Apply month and year filters\n2. Submit","Only executions for that period are shown","Month/year filters work correctly","Medium","Not Run"],
    ["TC-EM-005","Exec Monitor","Execution Details","Clicking an execution shows details","Execution list loaded","1. Click on an execution row","Detail panel/modal opens with configuration, status, results, metadata","Execution details view shows all relevant information for the selected execution","Critical","Not Run"],
    ["TC-EM-006","Exec Monitor","KPIs","KPI metrics load for selected execution","Execution selected","1. Select an execution\n2. View KPI section","KPI metrics (FTE, CPH, capacity, gap) are shown for that execution","KPI metrics are calculated and displayed per execution","High","Not Run"],
    ["TC-EM-007","Exec Monitor","Download","Download execution summary report","Execution selected","1. Select an execution\n2. Click Download Summary","Excel file with summary report is downloaded","Execution summary report downloads as a valid Excel file","Medium","Not Run"],
    ["TC-EM-008","Exec Monitor","Download","Download execution error log","Failed execution selected","1. Select a FAILED execution\n2. Click Download Error Log","Error log file is downloaded","Error log download is available for failed executions","Medium","Not Run"],
    ["TC-EM-009","Exec Monitor","Health","Health check endpoint responds","App running","1. Call /centene_forecasting/api/execution-monitoring/health/","200 response with system health status","Health check endpoint confirms system is operational","Low","Not Run"],
    ["TC-EM-010","Exec Monitor","Auto-refresh","KPI hero section auto-refreshes","Page loaded","1. Leave the page open for 5+ seconds\n2. Observe KPI section","KPI cards refresh automatically every 5 seconds","Auto-refresh interval (5000ms) keeps KPIs current without manual reload","High","Not Run"],
    ["TC-EM-011","Exec Monitor","Validation","Invalid execution ID returns error","User manipulates URL","1. Call details API with execution_id=abc","400 error returned","Non-numeric/invalid execution IDs are rejected","Medium","Not Run"],
    ["TC-EM-012","Exec Monitor","Pagination","Lazy load more executions","More than 100 executions exist","1. Scroll to bottom of execution list\n2. Trigger load more","Next 100 records are appended to the list","Pagination loads additional records on demand","Low","Not Run"],
]

# ── EDIT VIEW ─────────────────────────────────────────────────────────────────
EV_ROWS = [
    ["TC-EV-001","Edit View","Page Load","Edit view page loads with 3 tabs","User logged in with 'view' permission","1. Navigate to /centene_forecasting/edit-view/","Page loads; 3 tabs visible: Bench Allocation, Target CPH, Forecast Reallocation","Edit view renders with all three allocation management tabs","High","Not Run"],
    ["TC-EV-002","Edit View","Bench Alloc","Bench allocation preview loads","User has 'edit' permission; allocation reports available","1. Open Bench Allocation tab\n2. Select a report month/year\n3. Click Preview","Preview table appears with proposed allocation changes","Bench allocation preview shows projected changes before committing","Critical","Not Run"],
    ["TC-EV-003","Edit View","Bench Alloc","Bench allocation update submits","Preview shown; user approves","1. Preview bench allocation\n2. Click Apply/Update","Success message; change logged in history","Confirming bench allocation preview persists the change and logs it","Critical","Not Run"],
    ["TC-EV-004","Edit View","Bench Alloc","Missing month/year shows validation error","User on bench allocation tab","1. Click Preview without selecting month/year","Validation error shown","Month and year are required to generate a bench allocation preview","High","Not Run"],
    ["TC-EV-005","Edit View","Target CPH","CPH data table loads","User on Target CPH tab","1. Open Target CPH tab","Table with current CPH values per LOB/case type loads","Target CPH tab loads current CPH configuration data","High","Not Run"],
    ["TC-EV-006","Edit View","Target CPH","CPH change preview shows impact","CPH tab loaded; data visible","1. Edit a CPH value for a LOB/case type\n2. Click Preview","Impact analysis card shown: estimated FTE change, capacity change","CPH preview shows the downstream impact before user commits","Critical","Not Run"],
    ["TC-EV-007","Edit View","Target CPH","CPH update persists","Preview shown","1. Preview CPH change\n2. Confirm update","CPH value updated; success notification; history logged","CPH update persists the new value and records the change","Critical","Not Run"],
    ["TC-EV-008","Edit View","Target CPH","CPH value outside range rejected","User enters CPH value","1. Enter CPH = 0 or CPH = 200\n2. Click Preview","Validation error: CPH must be between 0.1 and 100.0","Out-of-range CPH values are rejected before processing","Critical","Not Run"],
    ["TC-EV-009","Edit View","Target CPH","CPH = 0.1 (min boundary) accepted","User enters boundary value","1. Enter CPH = 0.1\n2. Preview","Preview generated without validation error","Minimum boundary value (0.1) is accepted","Medium","Not Run"],
    ["TC-EV-010","Edit View","Target CPH","CPH = 100 (max boundary) accepted","User enters boundary value","1. Enter CPH = 100\n2. Preview","Preview generated without validation error","Maximum boundary value (100) is accepted","Medium","Not Run"],
    ["TC-EV-011","Edit View","Forecast Realloc","Reallocation filters load","User on Forecast Reallocation tab","1. Open Forecast Reallocation tab","Filter dropdowns populated (LOB, market, period)","Reallocation filters load correctly","High","Not Run"],
    ["TC-EV-012","Edit View","Forecast Realloc","Reallocation data table loads","Reallocation filters applied","1. Apply filters\n2. Load data","Table shows current allocation by period","Reallocation data is displayed for the selected filters","High","Not Run"],
    ["TC-EV-013","Edit View","Forecast Realloc","Reallocation preview shows diff","Data loaded; user modifies values","1. Change allocation values\n2. Click Preview","Diff preview shows before/after values","Reallocation preview shows what will change before committing","Critical","Not Run"],
    ["TC-EV-014","Edit View","Forecast Realloc","Reallocation update applies","Preview shown","1. Confirm reallocation","Change applied; history logged","Reallocation update persists the changes","Critical","Not Run"],
    ["TC-EV-015","Edit View","History Log","History log shows past changes","Changes have been made","1. Navigate to any edit tab\n2. View history section","Table with change history: timestamp, user, type, before, after","History log captures all edit operations with full audit trail","High","Not Run"],
    ["TC-EV-016","Edit View","History Log","History log can be downloaded","History exists","1. Click Download History","Excel file with change history downloaded","History log is exportable as Excel for audit purposes","Medium","Not Run"],
    ["TC-EV-017","Edit View","History Log","History filtered by change type","History log open","1. Select a change type filter\n2. Apply","Only records of that type shown","Change type filter correctly scopes history display","Medium","Not Run"],
    ["TC-EV-018","Edit View","Permissions","Viewer cannot submit edits","User with only 'view' permission","1. Try to click Preview/Apply on any edit tab","Access denied (403) or button disabled","Edit operations require 'edit' permission; view-only users are blocked","Critical","Not Run"],
    ["TC-EV-019","Edit View","Alloc Reports","Allocation report dropdown populates","Reports available from backend","1. Open edit view\n2. Observe allocation report dropdown","Dropdown shows available report options","Allocation reports dropdown correctly populates from API","High","Not Run"],
    ["TC-EV-020","Edit View","Target CPH","Negative CPH value rejected","User enters negative number","1. Enter CPH = -5\n2. Click Preview","Validation error shown","Negative CPH values are rejected","High","Not Run"],
]

# ── CONFIG ────────────────────────────────────────────────────────────────────
CFG_ROWS = [
    ["TC-CFG-001","Configuration","Page Load","Configuration page loads with 2 tabs","User logged in with 'view' permission","1. Navigate to /centene_forecasting/configuration/","Page loads; 2 tabs: Month Config, Target CPH","Configuration page renders both tabs","High","Not Run"],
    ["TC-CFG-002","Configuration","Month Config","Month config list loads","Configurations exist","1. Open Month Config tab","Table shows all month configurations","Month configuration list loads and displays all records","High","Not Run"],
    ["TC-CFG-003","Configuration","Month Config","Create single month config","User has 'add' permission","1. Click Add New\n2. Fill month=January, year=2025, work_type=Domestic, fte_count, hours\n3. Submit","Config saved; appears in list","A new month configuration can be created with valid data","Critical","Not Run"],
    ["TC-CFG-004","Configuration","Month Config","Create config with invalid month name","User creating month config","1. Enter month=13 or month=InvalidMonth\n2. Submit","Validation error: month must be January-December","Month name validation rejects invalid values","High","Not Run"],
    ["TC-CFG-005","Configuration","Month Config","Create config with year out of range","User creating month config","1. Enter year=1999 or year=2101\n2. Submit","Validation error: year must be 2020-2100","Year range validation rejects out-of-bounds values","High","Not Run"],
    ["TC-CFG-006","Configuration","Month Config","Create config with invalid work_type","User creating month config","1. Enter work_type=International\n2. Submit","Validation error: must be Domestic or Global","Work type enum validation rejects unknown values","High","Not Run"],
    ["TC-CFG-007","Configuration","Month Config","Bulk create month configs","User has 'add' permission","1. Prepare array of 3 month configs\n2. Submit bulk create API","All 3 configs saved; appear in list","Bulk create correctly saves multiple configurations atomically","Medium","Not Run"],
    ["TC-CFG-008","Configuration","Month Config","Update existing month config","Config exists; user has 'edit' permission","1. Select config\n2. Modify FTE count\n3. Save","Config updated; new value reflected in list","Updating a month configuration persists the change","High","Not Run"],
    ["TC-CFG-009","Configuration","Month Config","Delete month config","Config exists; user has 'delete' permission","1. Select config\n2. Click Delete\n3. Confirm","Config removed from list","Month configuration can be deleted by authorized users","Medium","Not Run"],
    ["TC-CFG-010","Configuration","Month Config","Validate config before save","User filling form","1. Click Validate\n2. Fix any errors\n3. Save","Validation runs inline; errors clearly highlighted","Validation endpoint can be called separately to check a config before committing","Low","Not Run"],
    ["TC-CFG-011","Configuration","Month Config","Filter month configs by work_type","Mixed configs exist","1. Apply work_type=Domestic filter","Only Domestic configs shown","Filter by work type narrows the configuration list","Medium","Not Run"],
    ["TC-CFG-012","Configuration","Target CPH","Target CPH list loads","CPH configs exist","1. Open Target CPH tab","Table shows all CPH configurations","Target CPH list loads and displays all records","High","Not Run"],
    ["TC-CFG-013","Configuration","Target CPH","Create target CPH config","User has 'add' permission","1. Click Add New CPH\n2. Fill main_lob, case_type, target_cph=10\n3. Submit","CPH config saved; appears in list","A new target CPH configuration can be created","Critical","Not Run"],
    ["TC-CFG-014","Configuration","Target CPH","Bulk create CPH configs","User has 'add' permission","1. Prepare array of 5 CPH configs\n2. Submit bulk","All 5 saved","Bulk create for CPH configs works correctly","Medium","Not Run"],
    ["TC-CFG-015","Configuration","Target CPH","Update target CPH","Config exists; user has 'edit' permission","1. Edit target_cph value\n2. Save","Updated value saved","Target CPH update persists correctly","High","Not Run"],
    ["TC-CFG-016","Configuration","Target CPH","Delete target CPH config","Config exists; user has 'delete' permission","1. Delete a CPH config\n2. Confirm","Config removed","Target CPH config can be deleted","Medium","Not Run"],
    ["TC-CFG-017","Configuration","Target CPH","Distinct LOBs API returns unique values","CPH data exists","1. Call /centene_forecasting/api/configuration/target-cph/distinct/main-lobs/","List of unique LOB names returned","Distinct LOBs API returns deduplicated LOB list for dropdowns","Low","Not Run"],
    ["TC-CFG-018","Configuration","Target CPH","Distinct case types API returns unique values","CPH data exists","1. Call /centene_forecasting/api/configuration/target-cph/distinct/case-types/","List of unique case type names returned","Distinct case types API returns deduplicated list","Low","Not Run"],
]

# ── CHAT ──────────────────────────────────────────────────────────────────────
CHAT_ROWS = [
    ["TC-CHAT-001","Chat","Connection","WebSocket connects successfully","User logged in","1. Open chat interface","WebSocket connection established; welcome message shown","Chat WebSocket connection is established when the chat page opens","Critical","Not Run"],
    ["TC-CHAT-002","Chat","Connection","Unauthenticated WebSocket connection rejected","No user session","1. Attempt to connect to WebSocket without session","Connection refused or immediate disconnect","Chat WebSocket requires valid authentication","Critical","Not Run"],
    ["TC-CHAT-003","Chat","Message","Send a text message and receive response","WebSocket connected","1. Type a message\n2. Press Send","Assistant responds within a reasonable time","LLM processes user messages and sends back a response","Critical","Not Run"],
    ["TC-CHAT-004","Chat","New Conversation","Start a new conversation","User in active chat","1. Click New Conversation","Current conversation ends; new conversation starts fresh with no prior context","Starting a new conversation resets context and history","High","Not Run"],
    ["TC-CHAT-005","Chat","Forecast Query","Query forecast data by month/year","Forecast data exists","1. Type: 'Show me forecast for January 2025'","Forecast data table is displayed for January 2025","Agent correctly interprets forecast query, fetches and displays data","Critical","Not Run"],
    ["TC-CHAT-006","Chat","Forecast Query","Filter update via chat","Chat context active","1. Type: 'Show data for Amisys platform'","Platform filter is applied; data refreshes for Amisys","Agent updates active filters based on natural language instructions","High","Not Run"],
    ["TC-CHAT-007","Chat","Available Reports","List available reports","Reports exist","1. Type: 'What reports are available?'","List of available months/reports shown","Agent correctly surfaces available forecast reports","High","Not Run"],
    ["TC-CHAT-008","Chat","FTE Details","Get FTE details","Context set with forecast data","1. Type: 'Show FTE details for this data'","FTE breakdown card is displayed","FTE details tool generates a structured breakdown card","High","Not Run"],
    ["TC-CHAT-009","Chat","CPH Update","Preview CPH change","Context has selected LOB/case type","1. Type: 'Change CPH to 15 for Amisys Medicare'","CPH preview card shown with impact analysis and Confirm button","CPH change preview shows impact analysis before committing","Critical","Not Run"],
    ["TC-CHAT-010","Chat","CPH Update","Confirm CPH change","CPH preview shown","1. Click Confirm Change button","CPH updated; success message","Confirming the CPH preview applies the change via confirm_cph_update WS message","Critical","Not Run"],
    ["TC-CHAT-011","Chat","CPH Update","Invalid CPH value rejected in chat","User asks for invalid CPH","1. Type: 'Set CPH to 0'","Error message: CPH must be between 0.1 and 100.0","CPH validation applies in chat context; invalid values produce a friendly error","High","Not Run"],
    ["TC-CHAT-012","Chat","Ramp","Trigger ramp calculation","Context has month/year","1. Type: 'Set up a ramp for January 2025'","Ramp trigger card shown with Configure Ramp button","setup_ramp_calculation tool generates the ramp entry UI","High","Not Run"],
    ["TC-CHAT-013","Chat","Ramp","Complete ramp modal form","Ramp trigger card shown","1. Click Configure Ramp\n2. Fill week percentages in modal\n3. Click Add","Ramp data captured; confirmation card shown","Ramp modal form captures weekly distribution and advances to confirmation","High","Not Run"],
    ["TC-CHAT-014","Chat","Ramp","Confirm and apply ramp","Ramp confirmation card shown","1. Click Yes Proceed\n2. Review preview\n3. Click Confirm Apply","Ramp applied; diff preview then result card shown","Full ramp flow completes from configuration to application","Critical","Not Run"],
    ["TC-CHAT-015","Chat","Ramp Campaign","Trigger ramp campaign","User in chat","1. Type: 'Set up ramp for multiple LOBs'","Ramp campaign entry card shown with Open Campaign Manager button","setup_ramp_campaign tool generates the campaign entry UI","High","Not Run"],
    ["TC-CHAT-016","Chat","Ramp Campaign","Add ramps in campaign manager","Campaign modal open (Stage view)","1. Click + Add New Ramp\n2. Select LOB and months\n3. Fill week data\n4. Click Add to Campaign","Row appears in staging table","Ramp campaign staging correctly accumulates multiple ramps","High","Not Run"],
    ["TC-CHAT-017","Chat","Ramp Campaign","Submit and preview campaign","Ramps staged","1. Click Submit All Ramps","Modal transitions to Preview view with all ramp diffs","Campaign submission generates a bulk preview across all staged ramps","Critical","Not Run"],
    ["TC-CHAT-018","Chat","Ramp Campaign","Confirm apply campaign","Campaign preview shown","1. Click Confirm Apply All","All ramps applied; modal transitions to Result view","Bulk campaign application applies all ramps in parallel","Critical","Not Run"],
    ["TC-CHAT-019","Chat","Bulk Ramp","View applied ramps","Ramps exist for a period","1. Type: 'Show me applied ramps for January 2025'","Ramp summary card with Show Data button","get_applied_ramp tool correctly fetches and displays existing ramps","High","Not Run"],
    ["TC-CHAT-020","Chat","Context","Clear context command","Active context exists","1. Type: 'Clear context' or 'Start over'","Context cleared; assistant acknowledges reset","Context can be explicitly cleared via natural language","Medium","Not Run"],
    ["TC-CHAT-021","Chat","Context","Context persists across messages","User set filters earlier in conversation","1. Set month/year via earlier message\n2. Send follow-up query without repeating filters","Follow-up query uses the established context","Conversation context is preserved and reused across messages in the same session","High","Not Run"],
    ["TC-CHAT-022","Chat","Download","Download ramp data as Excel","Ramp data exists","1. Navigate to /centene_forecasting/chat/download-ramp-excel/","Excel file with ramp data is downloaded","Ramp Excel download endpoint returns a valid file","Medium","Not Run"],
]

# ── CACHE ─────────────────────────────────────────────────────────────────────
CACHE_ROWS = [
    ["TC-CACHE-001","Cache","Stats","Cache stats API returns data","App running","1. Call /centene_forecasting/api/cache/stats/","JSON with cache backend info, TTLs, hit/miss rates","Cache stats endpoint provides visibility into cache health","Medium","Not Run"],
    ["TC-CACHE-002","Cache","Inspect","Inspect specific cache key","Data cached (e.g. after loading data view)","1. Call /centene_forecasting/api/cache/inspect/?key=forecast:2025:1","Returns key exists, type, size, preview","Cache inspection endpoint correctly reports on a specific cache key","Low","Not Run"],
    ["TC-CACHE-003","Cache","Clear","Clear forecast cache","Forecast data cached","1. POST to /centene_forecasting/api/cache/clear/forecast/","200 response; subsequent data view calls hit API (not cache)","Forecast cache clear removes only forecast-related keys","High","Not Run"],
    ["TC-CACHE-004","Cache","Clear","Clear roster cache","Roster data cached","1. POST to /centene_forecasting/api/cache/clear/roster/","200 response; roster data freshly fetched","Roster cache clear removes only roster-related keys","High","Not Run"],
    ["TC-CACHE-005","Cache","Clear","Clear cascade cache","Filter cascade results cached","1. POST to /centene_forecasting/api/cache/clear/cascade/","200 response; next filter cascade hits API","Cascade cache clear removes filter dropdown cached responses","Medium","Not Run"],
    ["TC-CACHE-006","Cache","Clear","Clear all caches","Multiple caches populated","1. POST to /centene_forecasting/api/cache/clear/all/","200 response; all subsequent requests hit the API","Clear all caches empties every cache type","Medium","Not Run"],
    ["TC-CACHE-007","Cache","Config","Cache config endpoint","App running","1. Call /centene_forecasting/api/cache/config/","Returns current cache configuration (backend, TTLs)","Cache config endpoint reflects the settings.py cache configuration","Low","Not Run"],
    ["TC-CACHE-008","Cache","TTL","Cascade filter cache expires after 300s","Data cached","1. Load filter cascade\n2. Wait 301 seconds\n3. Reload filter","API is called again (cache miss)","Cascade filter cache expires after 5 minutes per TTL configuration","Critical","Not Run"],
]

# ── Build all domain sheets ────────────────────────────────────────────────────
sheets_data = [
    ("Auth",        AUTH_ROWS),
    ("Upload",      UPLOAD_ROWS),
    ("DataView",    DV_ROWS),
    ("ManagerView", MV_ROWS),
    ("ExecMonitor", EM_ROWS),
    ("EditView",    EV_ROWS),
    ("Config",      CFG_ROWS),
    ("Chat",        CHAT_ROWS),
    ("Cache",       CACHE_ROWS),
]

for name, rows in sheets_data:
    ws = make_sheet(name)
    write_rows(ws, rows)

# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Summary", 0)
ws_sum.sheet_properties.tabColor = TAB_COLORS["Summary"]

sum_cols = ["Domain", "Sheet", "Total TCs", "Critical", "High", "Medium", "Low"]
sum_widths = [32, 16, 14, 12, 12, 12, 12]

hf = Font(name="Arial", bold=True, color=HDR_FG, size=11)
hfill = PatternFill("solid", fgColor=HDR_BG)
ha = Alignment(horizontal="center", vertical="center")

for i, (col, w) in enumerate(zip(sum_cols, sum_widths), start=1):
    c = ws_sum.cell(row=1, column=i, value=col)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin_border()
    ws_sum.column_dimensions[get_column_letter(i)].width = w

ws_sum.row_dimensions[1].height = 28
ws_sum.freeze_panes = "A2"

SUMMARY_DATA = [
    ("Authentication & Session",    "Auth",        12, 4, 5, 2, 1),
    ("Data Upload & File Mgmt",     "Upload",       14, 5, 5, 3, 1),
    ("Data View & Cascade Filters", "DataView",     18, 4, 7, 5, 2),
    ("Manager View Dashboard",      "ManagerView",  14, 3, 6, 3, 2),
    ("Execution Monitoring",        "ExecMonitor",  12, 2, 5, 3, 2),
    ("Edit View (Alloc & CPH)",     "EditView",     20, 6, 8, 4, 2),
    ("Configuration Management",   "Config",       18, 4, 7, 5, 2),
    ("Chat Assistant (LLM)",        "Chat",         22, 5, 9, 5, 3),
    ("Cache Management",           "Cache",         8, 1, 3, 3, 1),
]

pri_col_map = {4: COL_CRIT, 5: COL_HIGH, 6: COL_MED, 7: COL_LOW}

for r, row in enumerate(SUMMARY_DATA, start=2):
    is_alt = (r % 2 == 0)
    bg = ROW_ALT if is_alt else "FFFFFF"
    for c, val in enumerate(row, start=1):
        cell = ws_sum.cell(row=r, column=c, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left",
                                   vertical="center")
        color = pri_col_map.get(c, "000000")
        cell.font = Font(name="Arial", size=10,
                         bold=(c in pri_col_map),
                         color=color if c in pri_col_map else "000000")
        cell.fill = PatternFill("solid", fgColor=bg)
    ws_sum.row_dimensions[r].height = 22

# Totals row
tr = len(SUMMARY_DATA) + 2
total_label = ws_sum.cell(row=tr, column=1, value="TOTAL")
total_label.font = Font(name="Arial", bold=True, size=10)
total_label.fill = PatternFill("solid", fgColor="D9D9D9")
total_label.border = thin_border()

for c in range(2, 8):
    col_letter = get_column_letter(c)
    formula_cell = ws_sum.cell(row=tr, column=c,
                                value=f"=SUM({col_letter}2:{col_letter}{tr-1})")
    formula_cell.font = Font(name="Arial", bold=True, size=10,
                              color=pri_col_map.get(c, "000000"))
    formula_cell.fill = PatternFill("solid", fgColor="D9D9D9")
    formula_cell.border = thin_border()
    formula_cell.alignment = Alignment(horizontal="center")

ws_sum.row_dimensions[tr].height = 24

# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

out = "/Users/aswanthvishnu/Projects/Centene_Forecasting/docs/Manual_Test_Cases.xlsx"
wb.save(out)
print(f"Saved: {out}")
